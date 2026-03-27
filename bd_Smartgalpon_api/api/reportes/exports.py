import io
from django.db import connection
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from decimal import Decimal
from datetime import date


def _serialize(v):
    if isinstance(v, Decimal): return float(v)
    if isinstance(v, date):    return v.isoformat()
    return v


def get_reporte_data(lote_id):
    with connection.cursor() as cur:
        # FUNCTION → SELECT
        cur.execute("SELECT * FROM sp_resumen_ganancia_lote(%s)", [lote_id])
        cols    = [c[0] for c in cur.description]
        raw     = cur.fetchone() or []
        resumen = {k: _serialize(v) for k, v in zip(cols, raw)}

    with connection.cursor() as cur:
        cur.execute(
            "SELECT fecha, cantidad_muerta FROM registro_mortalidad "
            "WHERE lote_id = %s ORDER BY fecha", [lote_id])
        mortalidad = [(_serialize(f), c) for f, c in cur.fetchall()]

    with connection.cursor() as cur:
        cur.execute(
            "SELECT fecha, cantidad_huevos FROM registro_huevos "
            "WHERE lote_id = %s ORDER BY fecha", [lote_id])
        huevos = [(_serialize(f), c) for f, c in cur.fetchall()]

    with connection.cursor() as cur:
        cur.execute(
            "SELECT tipo, SUM(cantidad*precio) FROM insumos_ponedora "
            "WHERE lotes_id = %s GROUP BY tipo", [lote_id])
        insumos = [(t, _serialize(c)) for t, c in cur.fetchall()]

    return resumen, mortalidad, huevos, insumos


def exportar_pdf(lote_id):
    resumen, mortalidad, huevos, insumos = get_reporte_data(lote_id)
    buffer   = io.BytesIO()
    doc      = SimpleDocTemplate(buffer, pagesize=A4)
    styles   = getSampleStyleSheet()
    elements = []

    verde = colors.HexColor('#1D9E75')

    def tabla_con_estilo(data):
        t = Table(data, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), verde),
            ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
            ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0, 0), (-1, -1), 9),
            ('GRID',       (0, 0), (-1, -1), 0.4, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#F5F5F5')]),
        ]))
        return t

    elements.append(Paragraph(
        f"Reporte — {resumen.get('nombre', f'Lote {lote_id}')}",
        styles['Title']))
    elements.append(Spacer(1, 12))

    # Resumen financiero
    elements.append(Paragraph("Resumen financiero", styles['Heading2']))
    elements.append(tabla_con_estilo([
        ['Gallinas', 'Total huevos', 'Ingresos', 'Costos', 'Ganancia'],
        [
            str(resumen.get('cantidad_gallinas', '')),
            str(resumen.get('total_huevos', '')),
            f"${resumen.get('ingresos_totales', 0):.2f}",
            f"${resumen.get('costos_totales',  0):.2f}",
            f"${resumen.get('ganancia_total',  0):.2f}",
        ]
    ]))
    elements.append(Spacer(1, 12))

    # Mortalidad
    elements.append(Paragraph("Mortalidad", styles['Heading2']))
    elements.append(tabla_con_estilo(
        [['Fecha', 'Muertes']] + [[f, str(c)] for f, c in mortalidad]
        or [['Fecha', 'Muertes'], ['Sin registros', '']]))
    elements.append(Spacer(1, 12))

    # Huevos
    elements.append(Paragraph("Producción de huevos", styles['Heading2']))
    elements.append(tabla_con_estilo(
        [['Fecha', 'Huevos']] + [[f, str(c)] for f, c in huevos]
        or [['Fecha', 'Huevos'], ['Sin registros', '']]))
    elements.append(Spacer(1, 12))

    # Insumos
    elements.append(Paragraph("Insumos por tipo", styles['Heading2']))
    elements.append(tabla_con_estilo(
        [['Tipo', 'Costo total']] +
        [[t, f"${c:.2f}"] for t, c in insumos]
        or [['Tipo', 'Costo total'], ['Sin registros', '']]))

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = (
        f'attachment; filename="reporte_lote_{lote_id}.pdf"')
    return response


def exportar_excel(lote_id):
    resumen, mortalidad, huevos, insumos = get_reporte_data(lote_id)
    wb = openpyxl.Workbook()

    verde = 'FF1D9E75'

    def estilizar_header(ws, headers):
        from openpyxl.styles import PatternFill, Font, Alignment
        ws.append(headers)
        for cell in ws[1]:
            cell.fill      = PatternFill('solid', fgColor=verde)
            cell.font      = Font(color='FFFFFFFF', bold=True)
            cell.alignment = Alignment(horizontal='center')

    # Hoja resumen
    ws = wb.active
    ws.title = "Resumen"
    estilizar_header(ws, ['Campo', 'Valor'])
    for k, v in resumen.items():
        ws.append([k, v])

    # Hoja mortalidad
    ws2 = wb.create_sheet("Mortalidad")
    estilizar_header(ws2, ['Fecha', 'Cantidad muerta'])
    for row in mortalidad:
        ws2.append(list(row))

    # Hoja huevos
    ws3 = wb.create_sheet("Producción huevos")
    estilizar_header(ws3, ['Fecha', 'Cantidad huevos'])
    for row in huevos:
        ws3.append(list(row))

    # Hoja insumos
    ws4 = wb.create_sheet("Insumos")
    estilizar_header(ws4, ['Tipo', 'Costo total'])
    for row in insumos:
        ws4.append(list(row))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type=(
            'application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet'))
    response['Content-Disposition'] = (
        f'attachment; filename="reporte_lote_{lote_id}.xlsx"')
    return response